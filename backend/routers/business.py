import io
import re
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from database import get_session
from models import Expense, Item, Receipt, Sale

RECEIPTS_DIR = Path("/data/receipts")
router = APIRouter(prefix="/api/business", tags=["business"])

# ── Palette ──────────────────────────────────────────────────────────────────
BG_HEADER   = "1F3864"
BG_SECTION  = "2F75B6"
BG_ALT      = "F2F7FC"
BG_TOTAL    = "DEEAF1"
FG_WHITE    = "FFFFFF"
FG_DARK     = "1A1A2E"
FG_GREEN    = "375623"
FG_RED      = "C00000"
FG_SECTION  = "2F75B6"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)

def _font(bold=False, color=FG_DARK, size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _thin_border() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _header_row(ws, row: int, headers: list[str]):
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font  = _font(bold=True, color=FG_WHITE)
        c.fill  = _fill(BG_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()
    ws.row_dimensions[row].height = 22

def _money(ws, row: int, col: int, value: float, bold=False, color=FG_DARK):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = '$#,##0.00'
    c.font  = _font(bold=bold, color=color)
    c.border = _thin_border()
    return c

def _safe(s: str) -> str:
    return re.sub(r'[^\w\-]', '-', str(s))[:28].strip('-')


# ── Sheet builders ────────────────────────────────────────────────────────────

def _summary_sheet(ws, start: date, end: date, sales, expenses):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # ── Title banner
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "FlipTrack  ·  Business Report"
    c.font  = _font(bold=True, color=FG_WHITE, size=15)
    c.fill  = _fill(BG_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"Period: {start.strftime('%B %d, %Y')}  —  {end.strftime('%B %d, %Y')}"
    c.font  = _font(color=FG_WHITE, size=11)
    c.fill  = _fill(BG_SECTION)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    gross   = sum(s.sale_price    for s in sales)
    fees    = sum(s.platform_fees + s.shipping_cost for s in sales)
    net_s   = sum(s.net_profit    for s in sales)
    exp_tot = sum(e.amount        for e in expenses)
    net_inc = net_s - exp_tot

    def metric_row(r, label, value, is_total=False):
        bg = BG_TOTAL if is_total else "FFFFFF"
        lc = ws.cell(row=r, column=2, value=label)
        lc.font  = _font(bold=is_total, color=FG_DARK)
        lc.fill  = _fill(bg)
        lc.alignment = Alignment(indent=1)
        lc.border = _thin_border()
        vc = ws.cell(row=r, column=3, value=value)
        vc.number_format = '$#,##0.00'
        color = (FG_GREEN if value >= 0 else FG_RED) if value is not None else FG_DARK
        vc.font  = _font(bold=is_total, color=color)
        vc.fill  = _fill(bg)
        vc.border = _thin_border()
        ws.row_dimensions[r].height = 18

    def section_header(r, label):
        ws.merge_cells(f"B{r}:C{r}")
        c = ws.cell(row=r, column=2, value=label)
        c.font  = _font(bold=True, color=FG_WHITE, size=10)
        c.fill  = _fill(BG_SECTION)
        c.alignment = Alignment(indent=1)
        c.border = _thin_border()
        ws.row_dimensions[r].height = 18

    r = 4
    section_header(r, "INCOME"); r += 1
    metric_row(r, "Gross Revenue",          gross);     r += 1
    metric_row(r, "Platform & Shipping Fees", -fees);   r += 1
    metric_row(r, "Net Sales Profit",       net_s, is_total=True); r += 2

    section_header(r, "EXPENSES"); r += 1
    by_cat = defaultdict(float)
    for e in expenses:
        by_cat[e.category] += e.amount
    for cat, amt in sorted(by_cat.items()):
        metric_row(r, cat, -amt); r += 1
    metric_row(r, "Total Expenses", -exp_tot, is_total=True); r += 2

    section_header(r, "NET INCOME"); r += 1
    metric_row(r, "Net Income (after all expenses)", net_inc, is_total=True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 2


def _sales_sheet(ws, sales, items_map: dict, receipts_map: dict):
    ws.title = "Sales"
    ws.sheet_view.showGridLines = False

    headers = ["Date", "Item", "Sale Price", "Platform Fees", "Shipping", "Purchase Cost", "Net Profit", "Receipt Files"]
    widths  = [14, 32, 14, 16, 12, 16, 14, 44]
    _header_row(ws, 1, headers)

    for i, sale in enumerate(sorted(sales, key=lambda s: s.sold_date)):
        r   = i + 2
        bg  = BG_ALT if i % 2 else "FFFFFF"
        item = items_map.get(sale.item_id)
        cost = item.purchase_price if item else 0.0
        receipts = ", ".join(receipts_map.get(("item", sale.item_id), []))

        def cell(col, val):
            c = ws.cell(row=r, column=col, value=val)
            c.fill   = _fill(bg)
            c.border = _thin_border()
            return c

        c1 = cell(1, sale.sold_date); c1.number_format = "MMM DD, YYYY"; c1.alignment = Alignment(horizontal="center")
        cell(2, item.name if item else "Unknown").font = _font(color=FG_DARK)
        _money(ws, r, 3, sale.sale_price,    color=FG_DARK)
        _money(ws, r, 4, sale.platform_fees, color=FG_DARK)
        _money(ws, r, 5, sale.shipping_cost, color=FG_DARK)
        _money(ws, r, 6, cost,               color=FG_DARK)
        _money(ws, r, 7, sale.net_profit,    color=FG_GREEN if sale.net_profit >= 0 else FG_RED)
        cell(8, receipts).font = _font(color="666666", size=9)

    # Totals
    tr = len(sales) + 2
    def total_cell(col, val, fmt=None):
        c = ws.cell(row=tr, column=col, value=val)
        c.font   = _font(bold=True, color=FG_DARK)
        c.fill   = _fill(BG_TOTAL)
        c.border = _thin_border()
        if fmt: c.number_format = fmt
        return c

    total_cell(1, "")
    total_cell(2, "TOTALS")
    for col, attr in [(3, "sale_price"), (4, "platform_fees"), (5, "shipping_cost")]:
        total_cell(col, sum(getattr(s, attr) for s in sales), '$#,##0.00')
    total_cell(6, sum(items_map[s.item_id].purchase_price for s in sales if s.item_id in items_map), '$#,##0.00')
    c = total_cell(7, sum(s.net_profit for s in sales), '$#,##0.00')
    c.font = _font(bold=True, color=FG_GREEN if c.value >= 0 else FG_RED)
    total_cell(8, "")

    _set_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _expenses_sheet(ws, expenses, receipts_map: dict):
    ws.title = "Expenses"
    ws.sheet_view.showGridLines = False

    headers = ["Date", "Category", "Description", "Amount", "Receipt Files"]
    widths  = [14, 24, 42, 14, 44]
    _header_row(ws, 1, headers)

    by_cat = defaultdict(list)
    for e in sorted(expenses, key=lambda x: x.date):
        by_cat[e.category].append(e)

    row   = 2
    total = 0.0

    for cat in sorted(by_cat):
        # Category subheader stripe
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.fill   = _fill("EBF3FB")
            c.border = _thin_border()
        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=2, value=cat).font = _font(bold=True, color=FG_SECTION, size=10)
        sub = sum(e.amount for e in by_cat[cat])
        sc  = ws.cell(row=row, column=4, value=sub)
        sc.number_format = '$#,##0.00'
        sc.font   = _font(bold=True, color=FG_RED)
        sc.fill   = _fill("EBF3FB")
        sc.border = _thin_border()
        ws.row_dimensions[row].height = 16
        row += 1

        for i, exp in enumerate(by_cat[cat]):
            bg = BG_ALT if i % 2 else "FFFFFF"
            receipts = ", ".join(receipts_map.get(("expense", exp.id), []))

            def ecell(col, val):
                c = ws.cell(row=row, column=col, value=val)
                c.fill   = _fill(bg)
                c.border = _thin_border()
                return c

            dc = ecell(1, exp.date); dc.number_format = "MMM DD, YYYY"; dc.alignment = Alignment(horizontal="center")
            ecell(2, exp.category).font  = _font(color=FG_DARK)
            ecell(3, exp.description or "").font = _font(color="555555")
            ac = ecell(4, exp.amount); ac.number_format = '$#,##0.00'; ac.font = _font(color=FG_RED)
            ecell(5, receipts).font = _font(color="666666", size=9)
            total += exp.amount
            row += 1

    # Grand total
    for col in range(1, 6):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(BG_TOTAL)
        c.border = _thin_border()
    ws.cell(row=row, column=2, value="TOTAL EXPENSES").font = _font(bold=True, color=FG_DARK)
    tc = ws.cell(row=row, column=4, value=total)
    tc.number_format = '$#,##0.00'
    tc.font   = _font(bold=True, color=FG_RED)
    tc.fill   = _fill(BG_TOTAL)
    tc.border = _thin_border()

    _set_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/export/zip")
def export_zip(
    start: date = Query(...),
    end:   date = Query(...),
    session: Session = Depends(get_session),
):
    sales    = session.exec(select(Sale).where(Sale.sold_date >= start, Sale.sold_date <= end)).all()
    expenses = session.exec(select(Expense).where(Expense.date >= start, Expense.date <= end)).all()
    items_map = {i.id: i for i in session.exec(select(Item)).all()}

    item_ids    = list({s.item_id for s in sales})
    expense_ids = [e.id for e in expenses]

    raw_receipts: list[Receipt] = []
    if item_ids:
        raw_receipts += session.exec(
            select(Receipt).where(Receipt.entity_type == "item", Receipt.entity_id.in_(item_ids))
        ).all()
    if expense_ids:
        raw_receipts += session.exec(
            select(Receipt).where(Receipt.entity_type == "expense", Receipt.entity_id.in_(expense_ids))
        ).all()

    # Build receipt maps
    receipts_map: dict[tuple, list[str]] = defaultdict(list)
    disk_files: list[tuple[Path, str]] = []

    for r in raw_receipts:
        disk = RECEIPTS_DIR / r.filename
        if not disk.exists():
            continue
        if r.entity_type == "item":
            item = items_map.get(r.entity_id)
            label = _safe(item.name) if item else "item"
        else:
            exp = next((e for e in expenses if e.id == r.entity_id), None)
            label = _safe(exp.description or exp.category) if exp else "expense"
        ext      = Path(r.filename).suffix
        zip_name = f"receipts/{r.entity_type}_{r.entity_id}_{label}_{r.id}{ext}"
        receipts_map[(r.entity_type, r.entity_id)].append(zip_name)
        disk_files.append((disk, zip_name))

    # Build XLSX
    wb   = Workbook()
    ws1  = wb.active
    ws2  = wb.create_sheet()
    ws3  = wb.create_sheet()
    _summary_sheet(ws1, start, end, sales, expenses)
    _sales_sheet(ws2, sales, items_map, receipts_map)
    _expenses_sheet(ws3, expenses, receipts_map)

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)

    # Bundle into ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"FlipTrack_Report_{start}_{end}.xlsx", xlsx_buf.read())
        for disk_path, zip_path in disk_files:
            zf.write(disk_path, zip_path)

    zip_buf.seek(0)
    fname = f"fliptrack_cpa_{start}_{end}.zip"
    return StreamingResponse(
        iter([zip_buf.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/summary")
def get_summary(
    start: date = Query(...),
    end:   date = Query(...),
    session: Session = Depends(get_session),
):
    sales    = session.exec(select(Sale).where(Sale.sold_date >= start, Sale.sold_date <= end)).all()
    expenses = session.exec(select(Expense).where(Expense.date >= start, Expense.date <= end)).all()

    by_cat: dict[str, float] = defaultdict(float)
    for e in expenses:
        by_cat[e.category] += e.amount

    gross    = sum(s.sale_price  for s in sales)
    fees     = sum(s.platform_fees + s.shipping_cost for s in sales)
    net_s    = sum(s.net_profit  for s in sales)
    exp_tot  = sum(e.amount      for e in expenses)

    return {
        "period":               {"start": str(start), "end": str(end)},
        "sales_count":          len(sales),
        "gross_revenue":        round(gross,   2),
        "total_fees":           round(fees,    2),
        "net_sales_profit":     round(net_s,   2),
        "total_expenses":       round(exp_tot, 2),
        "net_income":           round(net_s - exp_tot, 2),
        "expenses_by_category": {k: round(v, 2) for k, v in sorted(by_cat.items())},
    }
